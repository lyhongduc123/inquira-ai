from __future__ import annotations

import argparse
import asyncio
import csv
import json
import re
import statistics
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

import litellm
from ragas.embeddings.base import embedding_factory
from ragas.metrics.collections import AnswerRelevancy, AnswerCorrectness, Faithfulness
from ragas.metrics import DiscreteMetric

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.core.config import settings


DEFAULT_DATASET_PATH = (
    Path(__file__).resolve().parent
    / "evals"
    / "datasets"
    / "litqa_answer_validations.csv"
)
DEFAULT_OUTPUT_PATH = (
    Path(__file__).resolve().parent
    / "evals"
    / "experiments"
    / "litqa"
    / f"ragas_eval_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
)


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _safe_float(value: Any) -> float | None:
    try:
        if value is None:
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _load_jsonl(path: Path, max_samples: int | None = None) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            payload = line.strip()
            if not payload:
                continue
            row = json.loads(payload)
            if not isinstance(row, dict):
                continue
            rows.append(row)
            if max_samples is not None and len(rows) >= max_samples:
                break
    return rows


def _decode_cell(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text:
        return ""
    if text.startswith("[") or text.startswith("{"):
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return text
    return text


def _load_csv(path: Path, max_samples: int | None = None) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            rows.append({key: _decode_cell(value) for key, value in row.items()})
            if max_samples is not None and len(rows) >= max_samples:
                break
    return rows


def _load_dataset(path: Path, max_samples: int | None = None) -> List[Dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        return _load_csv(path, max_samples=max_samples)
    return _load_jsonl(path, max_samples=max_samples)


def _dataset_path_from_resume_file(output_path: Path) -> Path | None:
    if output_path.suffix.lower() == ".csv" or not output_path.exists():
        return None
    try:
        payload = json.loads(output_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    input_meta = payload.get("input")
    if not isinstance(input_meta, dict):
        return None
    dataset_path = str(input_meta.get("dataset_path") or "").strip()
    return Path(dataset_path) if dataset_path else None


def _resolve_export_csv_path(export_csv: str, output_path: Path) -> Path:
    csv_path = Path(export_csv) if export_csv else output_path.with_suffix(".csv")
    if csv_path.exists() and csv_path.is_dir():
        return csv_path / f"{output_path.stem}.csv"
    if not csv_path.suffix:
        return csv_path / f"{output_path.stem}.csv"
    return csv_path


def _resolve_export_xlsx_path(export_xlsx: str, output_path: Path) -> Path:
    xlsx_path = Path(export_xlsx) if export_xlsx else output_path.with_suffix(".xlsx")
    if xlsx_path.exists() and xlsx_path.is_dir():
        return xlsx_path / f"{output_path.stem}.xlsx"
    if not xlsx_path.suffix:
        return xlsx_path / f"{output_path.stem}.xlsx"
    return xlsx_path


def _build_eval_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    def _normalize_list(value: Any) -> List[str]:
        if isinstance(value, list):
            return [str(v).strip() for v in value if str(v).strip()]
        if isinstance(value, str) and value.strip().startswith("["):
            decoded = _decode_cell(value)
            if isinstance(decoded, list):
                return [str(v).strip() for v in decoded if str(v).strip()]
        if isinstance(value, str) and value.strip():
            return [v.strip() for v in value.split("|") if v.strip()]
        return []

    def _normalize_contexts(value: Any) -> List[str]:
        if isinstance(value, list):
            return [str(v).strip() for v in value if str(v).strip()]
        if isinstance(value, str) and value.strip().startswith("["):
            decoded = _decode_cell(value)
            if isinstance(decoded, list):
                return [str(v).strip() for v in decoded if str(v).strip()]
        if isinstance(value, str) and value.strip():
            return [value.strip()]
        return []

    records: List[Dict[str, Any]] = [
        {
            "question": str(row.get("question") or row.get("query") or "").strip(),
            "answer": str(row.get("answer") or "").strip(),
            "contexts": _normalize_contexts(row.get("contexts") or row.get("context")),
            "ground_truth": str(
                row.get("ground_truth")
                or row.get("reference")
                or row.get("key_passage")
                or ""
            ).strip(),
            "key_passage": str(row.get("key_passage") or "").strip(),
            "paper_ids": _normalize_list(
                row.get("paper_ids") or row.get("paper_id")
            ),
            "ground_truth_paper_ids": _normalize_list(
                row.get("ground_truth_paper_ids") or row.get("ground_truth_paper_id")
            ),
            "doi": _normalize_list(row.get("doi")),
            "validation_id": row.get("validation_id"),
        }
        for row in rows
    ]
    records = [
        row
        for row in records
        if row["question"]
        and row["answer"]
        and isinstance(row["contexts"], list)
        and row["contexts"]
    ]
    return records


RAGAS_METRIC_NAMES: List[str] = [
    "faithfulness",
    "answer_relevancy",
    "answer_correctness",
]


def _unique_preserve_order(values: List[str]) -> List[str]:
    seen: set[str] = set()
    unique: List[str] = []
    for value in values:
        normalized = str(value).strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        unique.append(normalized)
    return unique


def _extract_citation_markers(answer: str) -> List[str]:
    return _unique_preserve_order(re.findall(r"\(cite:([^)]+)\)", answer or "", re.IGNORECASE))


def _extract_context_citation_ids(contexts: List[Any]) -> List[str]:
    ids: List[str] = []
    for context in contexts:
        ids.extend(_extract_citation_markers(str(context)))
    return _unique_preserve_order(ids)


def _citation_validity_score(row: Dict[str, Any]) -> float:
    cited = _extract_citation_markers(str(row.get("answer") or ""))
    if not cited:
        return 0.0

    contexts = row.get("contexts") or []
    contexts = contexts if isinstance(contexts, list) else [contexts]
    citable_index_count = len([ctx for ctx in contexts if str(ctx).strip()])

    valid_index_upper_bound = citable_index_count
    if valid_index_upper_bound == 0:
        valid_index_upper_bound = len(row.get("paper_ids") or [])

    valid_ids = set(
        _unique_preserve_order(
            [
                *[str(x).strip() for x in (row.get("paper_ids") or [])],
                *[str(x).strip() for x in (row.get("ground_truth_paper_ids") or [])],
                *_extract_context_citation_ids(contexts),
            ]
        )
    )

    matched = 0
    for marker in cited:
        citation_ref = marker.split("|", 1)[0].strip()
        if citation_ref.isdigit():
            idx = int(citation_ref)
            if 1 <= idx <= valid_index_upper_bound:
                matched += 1
            continue
        if citation_ref in valid_ids:
            matched += 1

    return matched / len(cited)


def _is_not_provided_answer(answer: str) -> bool:
    text = answer.lower()
    markers = [
        "don't know",
        "do not know",
        "i don't know",
        "i do not know",
        "not provided",
        "no evidence",
        "does not contain evidence",
        "insufficient evidence",
        "cannot be determined",
    ]
    return any(marker in text for marker in markers)


def _question_relevance_score(row: Dict[str, Any], embeddings: Any) -> float:
    try:
        nomic_query = f"search_query: {row['question']}"
        nomic_doc = f"search_document: {row['answer']}"

        qv = embeddings.embed_text(nomic_query)
        av = embeddings.embed_text(nomic_doc)
        if not qv or not av:
            return 0.0
        dot = sum(a * b for a, b in zip(qv, av))
        qn = sum(a * a for a in qv) ** 0.5
        an = sum(a * a for a in av) ** 0.5
        if qn == 0 or an == 0:
            return 0.0
        return max(0.0, min(1.0, dot / (qn * an)))
    except Exception:
        return 0.0


def _context_groundedness_score(row: Dict[str, Any], embeddings: Any) -> float:
    key_passage = str(row.get("key_passage") or row.get("ground_truth") or "").strip()
    if not key_passage:
        return 0.0

    try:
        nomic_key_doc = f"search_document: {key_passage}"
        nomic_answer_doc = f"search_document: {row['answer']}"

        kv = embeddings.embed_text(nomic_key_doc)
        av = embeddings.embed_text(nomic_answer_doc)
        if not kv or not av:
            return 0.0
            
        dot = sum(a * b for a, b in zip(kv, av))
        kn = sum(a * a for a in kv) ** 0.5
        an = sum(a * a for a in av) ** 0.5
        
        if kn == 0 or an == 0:
            return 0.0
            
        return max(0.0, min(1.0, dot / (kn * an)))
    except Exception:
        return 0.0


def _token_set(text: str) -> set[str]:
    return set(re.findall(r"[a-z0-9]+", (text or "").lower()))


def _overlap_ratio(a: str, b: str) -> float:
    ta = _token_set(a)
    tb = _token_set(b)
    if not ta:
        return 0.0
    return len(ta & tb) / len(ta)


def _key_passage_overlap_score(row: Dict[str, Any]) -> float:
    key_passage = str(row.get("key_passage") or row.get("ground_truth") or "").strip()
    if not key_passage:
        return 0.0
    contexts = [str(c).strip() for c in (row.get("contexts") or []) if str(c).strip()]
    return max((_overlap_ratio(key_passage, ctx) for ctx in contexts), default=0.0)


def _key_doi_retrieval_hit_score(row: Dict[str, Any]) -> float:
    retrieved_paper_ids = row.get("paper_ids") or []
    ground_truth_paper_ids = row.get("ground_truth_paper_ids") or []
    if not isinstance(retrieved_paper_ids, list) or not retrieved_paper_ids:
        return 0.0
    if not isinstance(ground_truth_paper_ids, list) or not ground_truth_paper_ids:
        return 0.0

    indexed_set = {str(x).strip() for x in ground_truth_paper_ids if str(x).strip()}
    retrieved_set = {str(x).strip() for x in retrieved_paper_ids if str(x).strip()}
    return 1.0 if indexed_set & retrieved_set else 0.0


class RagasEvaluate:
    def __init__(
        self,
        ragas_llm: Any,
        embeddings: Any,
    ) -> None:
        self.ragas_llm = ragas_llm
        self.embeddings = embeddings
        self.answer_relevancy_scorer = AnswerRelevancy(
            llm=ragas_llm, embeddings=embeddings
        )
        self.answer_correctness_scorer = AnswerCorrectness(
            llm=ragas_llm,
            embeddings=embeddings,
        )
        self.faithfulness_scorer = Faithfulness(llm=ragas_llm, embeddings=embeddings)
        self.discrete_metric = DiscreteMetric(
            name="answer_quality",
            allowed_values=["pass", "fail", "no_evidence"],
            prompt="Classify the quality of the reponse: {response} to the question: {user_input}. Return 'pass' if the answer is relevant, 'fail' if the answer is not relevant, and 'no_evidence' if the answer explicitly indicates that there is no evidence to provide an answer."
        )

def _extract_answer_quality_value(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"pass", "fail", "no_evidence"}:
            return normalized
        return None

    wrapped_value = getattr(value, "value", None)
    if wrapped_value is not None:
        wrapped = _extract_answer_quality_value(wrapped_value)
        if wrapped is not None:
            return wrapped

    if isinstance(value, dict):
        for candidate_key in ("answer_quality", "value", "label"):
            parsed = _extract_answer_quality_value(value.get(candidate_key))
            if parsed is not None:
                return parsed
        for item in value.values():
            parsed = _extract_answer_quality_value(item)
            if parsed is not None:
                return parsed

    if isinstance(value, (list, tuple)):
        for item in value:
            parsed = _extract_answer_quality_value(item)
            if parsed is not None:
                return parsed

    return None


async def _score_row_with_ragas(
    *,
    row: Dict[str, Any],
    scorer: RagasEvaluate,
) -> Dict[str, Any]:
    answer_relevancy_result, faithfulness_result, answer_correctness_result, answer_quality_result = await asyncio.gather(
        scorer.answer_relevancy_scorer.ascore(
            user_input=row["question"],
            response=row["answer"],
        ),
        scorer.faithfulness_scorer.ascore(
            user_input=row["question"],
            response=row["answer"],
            retrieved_contexts=row["contexts"],
        ),
        scorer.answer_correctness_scorer.ascore(
            user_input=row["question"],
            response=row["answer"],
            reference=f"ground_truth: {row.get('ground_truth')}, key_passage: {row.get('key_passage')}",
        ),
        scorer.discrete_metric.ascore(
            user_input=row["question"],
            response=row["answer"],
            llm=scorer.ragas_llm,
        ),
    )

    answer_relevancy = _extract_score_value(answer_relevancy_result, "answer_relevancy")
    faithfulness = _extract_score_value(faithfulness_result, "faithfulness")
    answer_correctness = _extract_score_value(
        answer_correctness_result, "answer_correctness"
    )
    answer_quality = _extract_answer_quality_value(answer_quality_result)

    return {
        "answer_relevancy": answer_relevancy,
        "faithfulness": faithfulness,
        "answer_correctness": answer_correctness,
        "answer_quality": answer_quality,
    }
        

def _extract_score_value(value: Any, metric_name: str | None = None) -> float | None:
    if value is None:
        return None

    direct = _safe_float(value)
    if direct is not None:
        return direct

    wrapped_value = getattr(value, "value", None)
    if wrapped_value is not None:
        wrapped = _extract_score_value(wrapped_value, metric_name)
        if wrapped is not None:
            return wrapped

    if isinstance(value, dict):
        if metric_name:
            metric_value = _extract_score_value(value.get(metric_name), metric_name)
            if metric_value is not None:
                return metric_value
        for item in value.values():
            parsed = _extract_score_value(item, metric_name)
            if parsed is not None:
                return parsed

    if isinstance(value, (list, tuple)):
        for item in value:
            parsed = _extract_score_value(item, metric_name)
            if parsed is not None:
                return parsed

    return None


def _build_ragas_llm(model: str, *, llm_max_tokens: int) -> tuple[Any, str, Any]:
    from openai import AsyncOpenAI
    from ragas.llms import llm_factory

    resolved_model = model.strip()
    if not resolved_model:
        raise ValueError("RAGAS judge model is empty")

    client: Any
    if resolved_model.startswith("openai/"):
        resolved_model = resolved_model.split("/", 1)[1]
        client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)
    elif resolved_model.startswith("gemini/"):
        resolved_model = resolved_model.split("/", 1)[1]
        client = AsyncOpenAI(
            api_key=settings.GEMINI_API_KEY,
            base_url="https://generativelanguage.googleapis.com/v1beta/openai/",
        )
    elif resolved_model.startswith("openrouter/"):
        resolved_model = resolved_model.split("/", 1)[1]
        client = AsyncOpenAI(
            api_key=settings.OPENROUTER_API_KEY,
            base_url="https://openrouter.ai/api/v1",
        )
    else:
        client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)

    ragas_llm = llm_factory(
        resolved_model, client=client
    )
    return ragas_llm, resolved_model, client
    
        
async def _update_answer_validations_from_results(
    *,
    per_sample: List[Dict[str, Any]],
) -> Dict[str, int]:
    try:
        from sqlalchemy import select

        from app.core.db.database import async_session, init_db
        from app.models.answer_vaidations import DBAnswerValidation
    except Exception:
        return {
            "attempted": 0,
            "updated": 0,
            "missing": 0,
            "invalid": 0,
            "skipped_no_id": len(per_sample),
        }

    def _to_int(value: Any) -> int | None:
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    id_by_idx: Dict[int, int] = {}
    skipped_no_id = 0
    invalid = 0

    for idx, row in enumerate(per_sample):
        raw_id = row.get("validation_id")
        validation_id = _to_int(raw_id)
        if validation_id is None:
            skipped_no_id += 1
            continue
        if validation_id <= 0:
            invalid += 1
            continue
        id_by_idx[idx] = validation_id

    if not id_by_idx:
        return {
            "attempted": 0,
            "updated": 0,
            "missing": 0,
            "invalid": invalid,
            "skipped_no_id": skipped_no_id,
        }

    await init_db()
    session = async_session()
    try:
        ids = sorted(set(id_by_idx.values()))
        rows = (
            (
                await session.execute(
                    select(DBAnswerValidation).where(DBAnswerValidation.id.in_(ids))
                )
            )
            .scalars()
            .all()
        )
        by_id = {r.id: r for r in rows}

        updated = 0
        missing = 0
        for idx, validation_id in id_by_idx.items():
            target = by_id.get(validation_id)
            if target is None:
                missing += 1
                continue

            row_data = per_sample[idx]
            faithfulness = _safe_float(row_data.get("faithfulness"))
            relevancy = _safe_float(row_data.get("answer_relevancy"))

            if faithfulness is not None:
                target.factual_accuracy_score = faithfulness
            if relevancy is not None:
                target.relevance_score = relevancy
            target.validated_at = datetime.now(timezone.utc)
            updated += 1

        await session.commit()
        return {
            "attempted": len(id_by_idx),
            "updated": updated,
            "missing": missing,
            "invalid": invalid,
            "skipped_no_id": skipped_no_id,
        }
    finally:
        await session.close()


def run_evaluation(
    *,
    dataset_path: Path,
    output_path: Path,
    model: str,
    embedding_provider: str,
    embedding_model: str,
    metrics_csv: str,
    metric_strictness: int,
    llm_max_tokens: int,
    max_samples: int | None,
    update_answer_validations: bool,
    resume: bool,
    max_concurrent: int,
    recompute_citation_validity: bool = False,
) -> Dict[str, Any]:
    rows = _load_dataset(dataset_path, max_samples=max_samples)
    if not rows:
        raise ValueError(f"No evaluation rows found in {dataset_path}")

    eval_rows = _build_eval_rows(rows)
    per_sample: List[Dict[str, Any]] = []
    if resume and output_path.exists():
        try:
            if output_path.suffix.lower() == ".csv":
                per_sample = _load_csv(output_path)
            else:
                existing_data = json.loads(output_path.read_text(encoding="utf-8"))
                per_sample = existing_data.get("per_sample", [])
            print(f"Resuming from {output_path}. Loaded {len(per_sample)} existing evaluations.")
        except Exception as e:
            print(f"Warning: Could not load resume file {output_path}: {e}")

    evaluated_ids = {
        str(s["validation_id"])
        for s in per_sample
        if s.get("validation_id") is not None
    }
    pending_rows = [
        r for r in eval_rows
        if str(r["validation_id"]) not in evaluated_ids
    ]

    if recompute_citation_validity and per_sample:
        rows_by_id = {
            str(row["validation_id"]): row
            for row in eval_rows
            if row.get("validation_id") is not None
        }
        recalculated = 0
        missing_source_rows = 0
        for sample in per_sample:
            validation_id = sample.get("validation_id")
            source_row = rows_by_id.get(str(validation_id))
            if source_row is None:
                missing_source_rows += 1
                continue
            sample["citation_validity"] = _citation_validity_score(source_row)
            recalculated += 1
        print(
            "Recomputed citation_validity for "
            f"{recalculated} resumed rows"
            + (f" ({missing_source_rows} missing source rows)" if missing_source_rows else "")
            + "."
        )

    metric_names = list(RAGAS_METRIC_NAMES)
    ragas_meta: Dict[str, Any] = {
        "enabled": True,
        "model": model,
        "resolved_model": model,
    }

    def _save_checkpoint(current_per_sample: List[Dict[str, Any]]) -> Dict[str, Any]:
        result_dict = {
            "citation_validity": [s.get("citation_validity") for s in current_per_sample],
            "question_relevance": [s.get("question_relevance") for s in current_per_sample],
            "context_groundedness": [s.get("context_groundedness") for s in current_per_sample],
            "key_passage_overlap": [s.get("key_passage_overlap") for s in current_per_sample],
            "key_doi_retrieval_hit": [s.get("key_doi_retrieval_hit") for s in current_per_sample],
            "answer_relevancy": [s.get("answer_relevancy") for s in current_per_sample],
            "faithfulness": [s.get("faithfulness") for s in current_per_sample],
            "answer_correctness": [s.get("answer_correctness") for s in current_per_sample],
            "answer_quality": [s.get("answer_quality") for s in current_per_sample],
        }

        no_evidence_count = sum(1 for row in current_per_sample if row.get("is_not_provided") is True)

        def _mean(name: str, *, skip_not_provided: bool = False) -> float | None:
            values: List[float] = []
            for idx, item in enumerate(result_dict.get(name, []) or []):
                if skip_not_provided and idx < len(current_per_sample):
                    if current_per_sample[idx].get("is_not_provided") is True:
                        continue
                parsed = _safe_float(item)
                if parsed is None:
                    continue
                values.append(parsed)
            return float(statistics.mean(values)) if values else None

        summary_checkpoint = {
            "timestamp": _utc_now_iso(),
            "input": {
                "dataset_path": str(dataset_path),
                "model": model,
                "embedding_provider": embedding_provider,
                "embedding_model": embedding_model,
                "metrics": metric_names,
                "metric_strictness": metric_strictness,
                "llm_max_tokens": llm_max_tokens,
                "max_samples": max_samples,
            },
            "counts": {
                "source_rows": len(rows),
                "evaluated_rows": len(current_per_sample),
                "no_evidence_rows": no_evidence_count,
            },
            "metrics_mean": {
                "citation_validity": _mean("citation_validity"),
                "question_relevance": _mean("question_relevance", skip_not_provided=True),
                "context_groundedness": _mean("context_groundedness", skip_not_provided=True),
                "key_passage_overlap": _mean("key_passage_overlap"),
                "key_doi_retrieval_hit": _mean("key_doi_retrieval_hit"),
                "answer_relevancy": _mean("answer_relevancy", skip_not_provided=True),
                "faithfulness": _mean("faithfulness", skip_not_provided=True),
                "answer_correctness": _mean("answer_correctness"),
            },
            "metrics_mean_including_no_evidence": {
                "question_relevance": _mean("question_relevance"),
                "context_groundedness": _mean("context_groundedness"),
                "answer_relevancy": _mean("answer_relevancy"),
                "faithfulness": _mean("faithfulness"),
            },
            "ragas": ragas_meta,
            "per_sample": current_per_sample,
            "raw": result_dict,
        }

        if output_path.suffix.lower() == ".csv":
            export_evaluation_to_csv(
                eval_results=summary_checkpoint,
                dataset_path=dataset_path,
                output_csv_path=output_path,
            )
        else:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(
                json.dumps(summary_checkpoint, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        return summary_checkpoint

    if pending_rows:
        print(f"Evaluating {len(pending_rows)} pending rows...")
        _ = litellm
        if embedding_provider.lower() == "ollama":
            resolved_embedding_model = (
                embedding_model
                if embedding_model.lower().startswith("ollama/")
                else f"ollama/{embedding_model}"
            )
            ragas_embeddings = embedding_factory(
                "litellm",
                model=resolved_embedding_model,
                api_base=settings.OLLAMA_BASE_URL,
            )
        else:
            ragas_embeddings = embedding_factory(embedding_provider, model=embedding_model)

        ragas_llm, resolved_model, client = _build_ragas_llm(
            model,
            llm_max_tokens=llm_max_tokens,
        )
        ragas_meta["resolved_model"] = resolved_model
        
        scorer = RagasEvaluate(
            ragas_llm=ragas_llm,
            embeddings=ragas_embeddings,
        )

        async def process_all_pending():
            semaphore = asyncio.Semaphore(max_concurrent)

            async def evaluate_and_format(row: Dict[str, Any]) -> Dict[str, Any]:
                async with semaphore:
                    ragas_score = await _score_row_with_ragas(row=row, scorer=scorer)
                    citation_validity = _citation_validity_score(row)
                    question_relevance = _question_relevance_score(row, ragas_embeddings)
                    context_groundedness = _context_groundedness_score(row, ragas_embeddings)
                    key_passage_overlap = _key_passage_overlap_score(row)
                    key_doi_retrieval_hit = _key_doi_retrieval_hit_score(row)
                    is_not_provided = _is_not_provided_answer(row["answer"])

                    return {
                        "validation_id": row.get("validation_id"),
                        "is_not_provided": is_not_provided,
                        "citation_validity": citation_validity,
                        "question_relevance": question_relevance,
                        "context_groundedness": context_groundedness,
                        "key_passage_overlap": key_passage_overlap,
                        "key_doi_retrieval_hit": key_doi_retrieval_hit,
                        "answer_relevancy": ragas_score.get("answer_relevancy"),
                        "faithfulness": ragas_score.get("faithfulness"),
                        "answer_correctness": ragas_score.get("answer_correctness"),
                        "answer_quality": ragas_score.get("answer_quality"),
                    }

            tasks = [asyncio.create_task(evaluate_and_format(row)) for row in pending_rows]
            for coro in asyncio.as_completed(tasks):
                completed_row = await coro
                per_sample.append(completed_row)
                _save_checkpoint(per_sample)
                print(f"Evaluated row validation_id={completed_row.get('validation_id')} | Total Completed: {len(per_sample)}/{len(eval_rows)}")

        asyncio.run(process_all_pending())

    else:
        print("All rows already evaluated. Recalculating summaries...")

    final_summary = _save_checkpoint(per_sample)

    if update_answer_validations:
        final_summary["db_update"] = asyncio.run(
            _update_answer_validations_from_results(per_sample=per_sample)
        )
    else:
        final_summary["db_update"] = {
            "attempted": 0,
            "updated": 0,
            "missing": 0,
            "invalid": 0,
            "skipped_no_id": 0,
        }

    if output_path.suffix.lower() == ".csv":
        export_evaluation_to_csv(
            eval_results=final_summary,
            dataset_path=dataset_path,
            output_csv_path=output_path,
        )
    else:
        output_path.write_text(
            json.dumps(final_summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    return final_summary


def export_evaluation_to_csv(
    *,
    eval_results: Dict[str, Any],
    dataset_path: Path,
    output_csv_path: Path,
) -> None:
    csv_rows = _build_export_rows(eval_results=eval_results, dataset_path=dataset_path)
    
    if csv_rows:
        fieldnames = csv_rows[0].keys()
        output_csv_path.parent.mkdir(parents=True, exist_ok=True)
        
        with output_csv_path.open("w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(csv_rows)
        
        print(f"Exported {len(csv_rows)} rows to CSV: {output_csv_path}")
    else:
        print("No rows to export to CSV")


def _build_export_rows(
    *,
    eval_results: Dict[str, Any],
    dataset_path: Path,
) -> List[Dict[str, Any]]:
    original_rows = _load_dataset(dataset_path)
    original_by_id: Dict[str, Dict[str, Any]] = {}
    for row in original_rows:
        vid = row.get("validation_id")
        if vid is not None:
            original_by_id[str(vid)] = row

    rows: List[Dict[str, Any]] = []
    for eval_row in eval_results.get("per_sample", []):
        validation_id = eval_row.get("validation_id")
        original_row = original_by_id.get(str(validation_id), {})
        contexts = original_row.get("contexts", []) or []
        if not isinstance(contexts, list):
            contexts = [str(contexts)]

        rows.append(
            {
                "validation_id": validation_id,
                "question": original_row.get("question", original_row.get("query", "")),
                "answer": original_row.get("answer", ""),
                "ground_truth": original_row.get("ground_truth", original_row.get("reference", "")),
                "key_passage": original_row.get("key_passage", ""),
                "paper_ids": json.dumps(original_row.get("paper_ids", []) or [], ensure_ascii=False),
                "ground_truth_paper_ids": json.dumps(original_row.get("ground_truth_paper_ids", []) or [], ensure_ascii=False),
                "context_count": len(contexts),
                "contexts": json.dumps(contexts, ensure_ascii=False),
                "is_not_provided": eval_row.get("is_not_provided", False),
                "citation_validity": eval_row.get("citation_validity"),
                "question_relevance": eval_row.get("question_relevance"),
                "context_groundedness": eval_row.get("context_groundedness"),
                "key_passage_overlap": eval_row.get("key_passage_overlap"),
                "key_doi_retrieval_hit": eval_row.get("key_doi_retrieval_hit"),
                "answer_relevancy": eval_row.get("answer_relevancy"),
                "faithfulness": eval_row.get("faithfulness"),
                "answer_correctness": eval_row.get("answer_correctness"),
                "answer_quality": eval_row.get("answer_quality"),
            }
        )
    return rows


def export_evaluation_to_xlsx(
    *,
    eval_results: Dict[str, Any],
    dataset_path: Path,
    output_xlsx_path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = _build_export_rows(eval_results=eval_results, dataset_path=dataset_path)
    output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "per_query"

    columns = [
        "validation_id",
        "question",
        "answer",
        "ground_truth",
        "key_passage",
        "answer_quality",
        "answer_correctness",
        "faithfulness",
        "answer_relevancy",
        "context_groundedness",
        "key_passage_overlap",
        "key_doi_retrieval_hit",
        "citation_validity",
        "is_not_provided",
        "context_count",
        "paper_ids",
        "ground_truth_paper_ids",
        "contexts",
    ]
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column) for column in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wrap_columns = {"question", "answer", "ground_truth", "key_passage", "contexts"}
    widths = {
        "validation_id": 14,
        "question": 58,
        "answer": 72,
        "ground_truth": 62,
        "key_passage": 62,
        "answer_quality": 16,
        "contexts": 90,
    }
    for idx, column in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = widths.get(column, 18)
        if column in wrap_columns:
            for cell in ws[letter]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            for cell in ws[letter]:
                cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("summary")
    summary.append(["metric", "value"])
    counts = eval_results.get("counts", {})
    metrics_mean = eval_results.get("metrics_mean", {})
    summary_rows = {
        "evaluated_rows": counts.get("evaluated_rows"),
        "no_evidence_rows": counts.get("no_evidence_rows"),
        **metrics_mean,
    }
    for key, value in summary_rows.items():
        summary.append([key, value])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 18

    wb.save(output_xlsx_path)
    print(f"Exported {len(rows)} rows to XLSX: {output_xlsx_path}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run RAGAS + heuristic evaluation over exported answer_validation dataset"
    )
    parser.add_argument(
        "--dataset-path",
        type=str,
        default=None,
        help=(
            "Path to exported CSV or JSONL dataset. When omitted with --resume, "
            "the dataset path recorded in the resume file is used if available."
        ),
    )
    parser.add_argument(
        "--model",
        type=str,
        default=(
            settings.LLM_MODEL[5] if settings.LLM_MODEL else "gemini/gemma-4-31b-it"
        ),
        help="LiteLLM model name for ragas llm_factory",
    )
    parser.add_argument(
        "--embedding-provider",
        type=str,
        default=settings.EMBEDDING_PROVIDER,
        help="Embedding provider (e.g. 'ollama', 'litellm', 'openai', 'google', 'huggingface')",
    )
    parser.add_argument(
        "--embedding-model",
        type=str,
        default=settings.OLLAMA_EMBEDDING_MODEL,
        help="Embedding model name (for Ollama this can be 'nomic-embed-text' or 'ollama/nomic-embed-text')",
    )
    parser.add_argument(
        "--metrics",
        type=str,
        default=",".join(RAGAS_METRIC_NAMES),
        help=(
            "Ignored. Metrics are fixed to faithfulness,answer_relevancy,answer_correctness "
            "plus heuristic metrics in code."
        ),
    )
    parser.add_argument(
        "--metric-strictness",
        type=int,
        default=1,
        help="Strictness used for judge metrics supporting majority voting (e.g., hallucination AspectCritic).",
    )
    parser.add_argument(
        "--llm-max-tokens",
        type=int,
        default=1024,
        help="Reserved evaluator max token setting (kept for CLI compatibility).",
    )
    parser.add_argument(
        "--max-samples",
        type=int,
        default=None,
        help="Optional cap on dataset rows",
    )
    parser.add_argument(
        "--max-concurrent",
        type=int,
        default=1,
        help="Maximum concurrent rows to process. Lower to 1 or 2 if hitting TimeoutErrors.",
    )
    parser.add_argument(
        "--output-file",
        type=str,
        default=str(DEFAULT_OUTPUT_PATH),
        help="Evaluation output path. CSV is the default; JSON remains supported for compatibility.",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Resume from output-file if it exists, skipping already evaluated rows.",
    )
    parser.add_argument(
        "--skip-db-update",
        action="store_true",
        help="Skip writing faithfulness/relevancy scores back to answer_validations table",
    )
    parser.add_argument(
        "--export-csv",
        type=str,
        default=None,
        help="Optional extra CSV export path.",
    )
    parser.add_argument(
        "--export-xlsx",
        type=str,
        default=None,
        help="Optional XLSX export path. Pass a directory to create <output-file-stem>.xlsx inside it.",
    )
    parser.add_argument(
        "--recompute-citation-validity",
        action="store_true",
        help=(
            "With --resume, recompute only citation_validity for existing rows before "
            "saving/exporting summaries."
        ),
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    output_path = Path(args.output_file)
    dataset_path = (
        Path(args.dataset_path)
        if args.dataset_path
        else (
            _dataset_path_from_resume_file(output_path)
            if args.resume
            else None
        )
    )
    dataset_path = dataset_path or DEFAULT_DATASET_PATH

    summary = run_evaluation(
        dataset_path=dataset_path,
        output_path=output_path,
        model=args.model,
        embedding_provider=args.embedding_provider,
        embedding_model=args.embedding_model,
        metrics_csv=args.metrics,
        metric_strictness=args.metric_strictness,
        llm_max_tokens=args.llm_max_tokens,
        max_samples=args.max_samples,
        update_answer_validations=not args.skip_db_update,
        resume=args.resume,
        max_concurrent=args.max_concurrent,
        recompute_citation_validity=args.recompute_citation_validity,
    )
    print(
        f"Evaluation complete. evaluated={summary['counts']['evaluated_rows']} "
        f"output={args.output_file}"
    )

    if args.export_csv or args.export_csv == "":
        export_evaluation_to_csv(
            eval_results=summary,
            dataset_path=dataset_path,
            output_csv_path=_resolve_export_csv_path(args.export_csv or "", output_path),
        )

    if args.export_xlsx or args.export_xlsx == "":
        export_evaluation_to_xlsx(
            eval_results=summary,
            dataset_path=dataset_path,
            output_xlsx_path=_resolve_export_xlsx_path(args.export_xlsx or "", output_path),
        )
    
if __name__ == "__main__":
    main()
